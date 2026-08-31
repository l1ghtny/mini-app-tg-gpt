from __future__ import annotations

import hashlib
import re
import unicodedata
import zipfile
from collections.abc import Iterable, Mapping, Sequence
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openai import AsyncOpenAI
from openpyxl import load_workbook


MAX_GENERATED_ARTIFACTS = 5
MAX_GENERATED_ARTIFACT_BYTES = 25_000_000
MAX_GENERATED_ARTIFACT_TOTAL_BYTES = 50_000_000

_SUPPORTED_OUTPUTS: dict[str, tuple[str, str, str | None]] = {
    ".csv": ("text/csv", "spreadsheet", "text"),
    ".docx": (
        "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        "document",
        None,
    ),
    ".jpeg": ("image/jpeg", "image", "image"),
    ".jpg": ("image/jpeg", "image", "image"),
    ".md": ("text/markdown", "document", "text"),
    ".pdf": ("application/pdf", "document", "pdf"),
    ".png": ("image/png", "image", "image"),
    ".pptx": (
        "application/vnd.openxmlformats-officedocument.presentationml.presentation",
        "presentation",
        None,
    ),
    ".txt": ("text/plain", "document", "text"),
    ".xlsx": (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "spreadsheet",
        None,
    ),
}


class GeneratedArtifactError(RuntimeError):
    pass


@dataclass(frozen=True)
class GeneratedArtifactReference:
    container_id: str
    file_id: str
    filename: str
    mime_type: str
    kind: str
    preview_kind: str | None


@dataclass(frozen=True)
class DownloadedGeneratedArtifact:
    reference: GeneratedArtifactReference
    path: Path
    size_bytes: int
    sha256: str


def build_generated_spreadsheet_preview(
    path: Path,
    *,
    goal: str | None,
    source_count: int,
) -> dict[str, object]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        column_count = min(int(worksheet.max_column or 0), 250_000)
        raw_row_count = min(int(worksheet.max_row or 0), 250_001)
        preview_width = min(column_count, 30)
        rows = worksheet.iter_rows(
            min_row=1,
            max_row=min(raw_row_count, 101),
            max_col=preview_width,
            values_only=True,
        )
        header = next(rows, ())
        columns = []
        for index in range(preview_width):
            raw_label = header[index] if index < len(header) else None
            label = str(raw_label).strip() if raw_label is not None else ""
            columns.append(
                {
                    "label": (label or f"Column {index + 1}")[:120],
                    "data_type": "text",
                }
            )
        preview_rows: list[list[str | int | float | bool | None]] = []
        for row in rows:
            values: list[str | int | float | bool | None] = []
            for index in range(preview_width):
                value = row[index] if index < len(row) else None
                if value is None or isinstance(value, (str, int, float, bool)):
                    normalized = value
                elif hasattr(value, "isoformat"):
                    normalized = value.isoformat()
                else:
                    normalized = str(value)
                if isinstance(normalized, str):
                    normalized = normalized[:500]
                values.append(normalized)
            preview_rows.append(values)
        for column_index, column in enumerate(columns):
            values = [
                row[column_index]
                for row in preview_rows
                if row[column_index] is not None
            ]
            if values and all(isinstance(value, bool) for value in values):
                column["data_type"] = "boolean"
            elif values and all(
                isinstance(value, (int, float)) and not isinstance(value, bool)
                for value in values
            ):
                column["data_type"] = "number"
        row_count = max(raw_row_count - 1, 0)
        rows_truncated = row_count > len(preview_rows)
        columns_truncated = column_count > preview_width
        warning_codes: list[str] = []
        if rows_truncated:
            warning_codes.append("preview_rows_truncated")
        if columns_truncated:
            warning_codes.append("preview_columns_truncated")
        return {
            "version": 1,
            "goal": goal[:4000] if goal else None,
            "row_count": row_count,
            "column_count": column_count,
            "source_count": source_count,
            "columns": columns,
            "rows": preview_rows,
            "rows_truncated": rows_truncated,
            "columns_truncated": columns_truncated,
            "warning_codes": warning_codes,
        }
    finally:
        workbook.close()


def _value(item: Any, name: str) -> Any:
    if isinstance(item, Mapping):
        return item.get(name)
    return getattr(item, name, None)


def _annotations(response: Any) -> Iterable[Any]:
    for item in _value(response, "output") or []:
        for content in _value(item, "content") or []:
            yield from _value(content, "annotations") or []


def _safe_filename(value: Any) -> str | None:
    if not isinstance(value, str):
        return None
    normalized = unicodedata.normalize("NFKC", Path(value).name).strip()
    normalized = "".join(
        character
        for character in normalized
        if character.isprintable() and character not in {"/", "\\", ":"}
    )
    if not normalized or normalized in {".", ".."}:
        return None
    suffix = Path(normalized).suffix.lower()
    if suffix not in _SUPPORTED_OUTPUTS:
        return None
    stem = Path(normalized).stem.strip(" .")[:180]
    return f"{stem or 'artifact'}{suffix}"


def generated_artifact_references(response: Any) -> list[GeneratedArtifactReference]:
    references: list[GeneratedArtifactReference] = []
    seen: set[tuple[str, str]] = set()
    for annotation in _annotations(response):
        if _value(annotation, "type") != "container_file_citation":
            continue
        container_id = _value(annotation, "container_id")
        file_id = _value(annotation, "file_id")
        filename = _safe_filename(_value(annotation, "filename"))
        if (
            not isinstance(container_id, str)
            or not container_id
            or not isinstance(file_id, str)
            or not file_id
            or filename is None
        ):
            continue
        identity = (container_id, file_id)
        if identity in seen:
            continue
        seen.add(identity)
        mime_type, kind, preview_kind = _SUPPORTED_OUTPUTS[Path(filename).suffix.lower()]
        references.append(
            GeneratedArtifactReference(
                container_id=container_id,
                file_id=file_id,
                filename=filename,
                mime_type=mime_type,
                kind=kind,
                preview_kind=preview_kind,
            )
        )
        if len(references) >= MAX_GENERATED_ARTIFACTS:
            break
    return references


def plan_expects_artifacts(expected_outputs: Sequence[Mapping[str, Any]]) -> bool:
    return any(output.get("kind") == "artifact" for output in expected_outputs)


def artifact_contract_error(
    response: Any,
    request_payload: Mapping[str, Any],
) -> str | None:
    approved_plan = request_payload.get("approved_plan")
    expected_outputs = (
        approved_plan.get("expected_outputs", [])
        if isinstance(approved_plan, Mapping)
        else []
    )
    artifact_outputs = [
        output
        for output in expected_outputs
        if isinstance(output, Mapping) and output.get("kind") == "artifact"
    ]
    if not artifact_outputs:
        return None
    references = generated_artifact_references(response)
    if not references:
        return "The approved deliverable requires a generated file, but none was cited."

    requested_groups: list[tuple[str, set[str]]] = []
    format_patterns = (
        (r"\bpdf\b", "PDF", {".pdf"}),
        (r"\b(?:docx|word document)\b", "Word document", {".docx"}),
        (r"\b(?:pptx|powerpoint|slide deck)\b", "presentation", {".pptx"}),
        (r"\b(?:xlsx|excel workbook)\b", "Excel workbook", {".xlsx"}),
        (r"\bcsv\b", "CSV", {".csv"}),
        (r"\bmarkdown\b", "Markdown", {".md"}),
        (r"\bpng\b", "PNG", {".png"}),
        (r"\b(?:jpe?g)\b", "JPEG", {".jpg", ".jpeg"}),
    )
    for output in artifact_outputs:
        label_text = str(output.get("label", ""))
        label_groups = [
            (label, extensions)
            for pattern, label, extensions in format_patterns
            if re.search(pattern, label_text, flags=re.IGNORECASE)
        ]
        if label_groups:
            requested_groups.append(label_groups[0])
            continue
        output_text = " ".join(
            str(value)
            for value in (
                output.get("description", ""),
                *(output.get("acceptance_criteria", []) or []),
            )
        )
        output_groups = [
            (label, extensions)
            for pattern, label, extensions in format_patterns
            if re.search(pattern, output_text, flags=re.IGNORECASE)
        ]
        if output_groups:
            requested_groups.append(output_groups[0])
    produced_extensions = {
        Path(reference.filename).suffix.lower() for reference in references
    }
    for label, extensions in requested_groups:
        if produced_extensions.isdisjoint(extensions):
            return (
                f"The approved deliverable requires a {label} file, but the response "
                f"only cited: {', '.join(sorted(produced_extensions))}."
            )
    return None


def _validate_ooxml(path: Path, suffix: str) -> None:
    required_prefix = {
        ".docx": "word/",
        ".pptx": "ppt/",
        ".xlsx": "xl/",
    }[suffix]
    try:
        with zipfile.ZipFile(path) as archive:
            names = archive.namelist()
    except (OSError, zipfile.BadZipFile) as exc:
        raise GeneratedArtifactError("generated Office file is invalid") from exc
    if "[Content_Types].xml" not in names or not any(
        name.startswith(required_prefix) for name in names
    ):
        raise GeneratedArtifactError("generated Office file has an invalid structure")


def _validate_content(path: Path, reference: GeneratedArtifactReference) -> None:
    suffix = Path(reference.filename).suffix.lower()
    prefix = path.read_bytes()[:8]
    if suffix == ".pdf" and not prefix.startswith(b"%PDF-"):
        raise GeneratedArtifactError("generated PDF has an invalid signature")
    if suffix == ".png" and prefix != b"\x89PNG\r\n\x1a\n":
        raise GeneratedArtifactError("generated PNG has an invalid signature")
    if suffix in {".jpg", ".jpeg"} and not prefix.startswith(b"\xff\xd8\xff"):
        raise GeneratedArtifactError("generated JPEG has an invalid signature")
    if suffix in {".docx", ".pptx", ".xlsx"}:
        _validate_ooxml(path, suffix)
    if suffix in {".csv", ".md", ".txt"}:
        try:
            path.read_text(encoding="utf-8-sig")
        except UnicodeDecodeError as exc:
            raise GeneratedArtifactError(
                "generated text artifact is not valid UTF-8"
            ) from exc


async def download_generated_artifact(
    client: AsyncOpenAI,
    reference: GeneratedArtifactReference,
    destination: Path,
) -> DownloadedGeneratedArtifact:
    size_bytes = 0
    digest = hashlib.sha256()
    stream_method = client.containers.files.content.with_streaming_response.retrieve
    async with stream_method(
        reference.file_id,
        container_id=reference.container_id,
    ) as response:
        with destination.open("wb") as output:
            async for chunk in response.iter_bytes():
                size_bytes += len(chunk)
                if size_bytes > MAX_GENERATED_ARTIFACT_BYTES:
                    raise GeneratedArtifactError("generated artifact is too large")
                digest.update(chunk)
                output.write(chunk)
    if size_bytes == 0:
        raise GeneratedArtifactError("generated artifact is empty")
    _validate_content(destination, reference)
    return DownloadedGeneratedArtifact(
        reference=reference,
        path=destination,
        size_bytes=size_bytes,
        sha256=digest.hexdigest(),
    )
