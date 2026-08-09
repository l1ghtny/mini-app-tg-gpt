from __future__ import annotations

import csv
import io
import json
import re
import uuid
from dataclasses import dataclass
from datetime import date, datetime, timezone
from decimal import Decimal
from pathlib import Path
from typing import Any, Iterable, Mapping

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

from app.services.document_source_validation import inspect_spreadsheet_source


_WHITESPACE = re.compile(r"\s+")
_NUMBER = re.compile(r"^-?(?:0|[1-9]\d*)(?:[.,]\d+)?$")
_FORMULA_PREFIXES = ("=", "+", "-", "@")
_MAX_OUTPUT_ROWS = 250_000
_PREVIEW_MAX_ROWS = 100
_PREVIEW_MAX_COLUMNS = 30
_PREVIEW_MAX_CELL_CHARS = 500
_PREVIEW_MAX_JSON_BYTES = 1_800_000
_PRICE_HEADER_HINTS = (
    "amount",
    "cost",
    "price",
    "revenue",
    "total",
    "итого",
    "сумм",
    "стоим",
    "цен",
)


class ComparisonInputError(ValueError):
    pass


@dataclass(frozen=True)
class SourceTable:
    document_id: uuid.UUID
    filename: str
    sheet_name: str
    headers: tuple[str, ...]
    rows: tuple[tuple[Any, ...], ...]
    first_data_row: int


@dataclass(frozen=True)
class ComparisonSource:
    document_id: uuid.UUID
    filename: str
    sheet_name: str
    row_start: int
    row_end: int
    output_row_start: int
    output_row_end: int


@dataclass(frozen=True)
class RenderedComparison:
    path: Path
    row_count: int
    column_count: int
    sources: tuple[ComparisonSource, ...]
    preview: dict[str, object]


@dataclass(frozen=True)
class ComparisonColumnSchema:
    canonical_headers: tuple[str, ...]
    source_headers: Mapping[tuple[uuid.UUID, str, str], str]

    def canonical_header(self, table: SourceTable, source_header: str) -> str:
        key = (table.document_id, table.sheet_name, source_header)
        try:
            return self.source_headers[key]
        except KeyError as exc:
            raise ComparisonInputError(
                "comparison schema does not map every source column"
            ) from exc


def _decode_csv(data: bytes) -> str:
    try:
        return data.decode("utf-8-sig")
    except UnicodeDecodeError:
        return data.decode("cp1251")


def _clean_header(value: Any, ordinal: int, seen: set[str]) -> str:
    text = _WHITESPACE.sub(" ", str(value or "").strip()) or f"Column {ordinal}"
    candidate = text
    suffix = 2
    while candidate.casefold() in seen:
        candidate = f"{text} ({suffix})"
        suffix += 1
    seen.add(candidate.casefold())
    return candidate


def _trim_row(values: Iterable[Any]) -> tuple[Any, ...]:
    row = list(values)
    while row and row[-1] is None:
        row.pop()
    return tuple(row)


def _table_from_rows(
    *,
    document_id: uuid.UUID,
    filename: str,
    sheet_name: str,
    rows: Iterable[Iterable[Any]],
) -> SourceTable | None:
    materialized = [_trim_row(row) for row in rows]
    materialized = [
        row for row in materialized if any(value not in (None, "") for value in row)
    ]
    if not materialized:
        return None
    seen: set[str] = set()
    headers = tuple(
        _clean_header(value, index, seen)
        for index, value in enumerate(materialized[0], start=1)
    )
    if not headers:
        return None
    data_rows = tuple(materialized[1:])
    return SourceTable(
        document_id=document_id,
        filename=filename,
        sheet_name=sheet_name,
        headers=headers,
        rows=data_rows,
        first_data_row=2,
    )


def load_source_tables(
    *,
    document_id: uuid.UUID,
    filename: str,
    path: Path,
) -> tuple[SourceTable, ...]:
    inspect_spreadsheet_source(str(path), filename)
    extension = path.suffix.lower()
    tables: list[SourceTable] = []
    if extension == ".csv":
        text = _decode_csv(path.read_bytes())
        try:
            dialect = csv.Sniffer().sniff(text[:8192], delimiters=",;\t|")
        except csv.Error:
            dialect = csv.excel
        table = _table_from_rows(
            document_id=document_id,
            filename=filename,
            sheet_name="CSV",
            rows=csv.reader(io.StringIO(text, newline=""), dialect=dialect),
        )
        if table:
            tables.append(table)
    elif extension == ".xlsx":
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            for worksheet in workbook.worksheets:
                table = _table_from_rows(
                    document_id=document_id,
                    filename=filename,
                    sheet_name=worksheet.title,
                    rows=worksheet.iter_rows(values_only=True),
                )
                if table:
                    tables.append(table)
        finally:
            workbook.close()
    else:
        raise ComparisonInputError(f"unsupported comparison source: {extension}")

    if not tables:
        raise ComparisonInputError(f"{filename} contains no tabular data")
    return tuple(tables)


def _safe_cell(value: Any) -> Any:
    if isinstance(value, datetime):
        if value.tzinfo is not None:
            return value.astimezone(timezone.utc).replace(tzinfo=None)
        return value
    if isinstance(value, str) and value.startswith(_FORMULA_PREFIXES):
        return f"'{value}"
    return value


def _non_blank(values: Iterable[Any]) -> list[Any]:
    return [value for value in values if value not in (None, "")]


def _numeric_string(value: str) -> bool:
    normalized = value.strip()
    if not _NUMBER.fullmatch(normalized):
        return False
    unsigned = normalized.removeprefix("-")
    integer_part = unsigned.replace(",", ".").split(".", 1)[0]
    return len(integer_part) == 1 or not integer_part.startswith("0")


def _parse_temporal_string(value: str) -> date | datetime | None:
    normalized = value.strip()
    try:
        if "T" in normalized or " " in normalized:
            parsed = datetime.fromisoformat(normalized.replace("Z", "+00:00"))
            if parsed.tzinfo is not None:
                return parsed.astimezone(timezone.utc).replace(tzinfo=None)
            return parsed
        return date.fromisoformat(normalized)
    except ValueError:
        return None


def _infer_data_type(values: Iterable[Any]) -> str:
    present = _non_blank(values)
    if not present:
        return "text"
    if all(isinstance(value, bool) for value in present):
        return "boolean"
    if all(isinstance(value, datetime) for value in present):
        return "datetime"
    if all(isinstance(value, date) for value in present):
        return "date"
    if all(
        isinstance(value, (int, float, Decimal)) and not isinstance(value, bool)
        for value in present
    ):
        return "number"
    if all(
        (isinstance(value, (int, float, Decimal)) and not isinstance(value, bool))
        or (isinstance(value, str) and _numeric_string(value))
        for value in present
    ):
        return "number"
    if all(isinstance(value, str) for value in present):
        temporal_values = [_parse_temporal_string(value) for value in present]
        if all(value is not None for value in temporal_values):
            if all(
                isinstance(value, date) and not isinstance(value, datetime)
                for value in temporal_values
            ):
                return "date"
            return "datetime"
    return "text"


def _coerce_cell(value: Any, data_type: str) -> Any:
    value = _safe_cell(value)
    if value in (None, ""):
        return None
    if data_type == "number" and isinstance(value, str):
        normalized = value.strip().replace(",", ".")
        if "." not in normalized:
            integer = int(normalized)
            if len(normalized.removeprefix("-")) <= 15:
                return integer
            return value
        return float(normalized)
    if data_type in {"date", "datetime"} and isinstance(value, str):
        return _parse_temporal_string(value) or value
    return value


def _preview_value(value: Any) -> str | int | float | bool | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, int):
        return value
    if isinstance(value, (float, Decimal)):
        return float(value)
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    text = str(value)
    if text.startswith("'") and text[1:].startswith(_FORMULA_PREFIXES):
        text = text[1:]
    return text[:_PREVIEW_MAX_CELL_CHARS]


def _column_format(header: str, data_type: str, currency: str | None) -> str | None:
    if data_type == "date":
        return "yyyy-mm-dd"
    if data_type == "datetime":
        return "yyyy-mm-dd hh:mm"
    if data_type != "number":
        return None
    if currency and any(hint in header.casefold() for hint in _PRICE_HEADER_HINTS):
        return f'#,##0.00 "{currency}"'
    return "#,##0.00"


def _column_width(header: str, values: Iterable[Any]) -> float:
    longest = len(header)
    for value in list(values)[:60]:
        if value is None:
            continue
        longest = max(
            longest, max((len(line) for line in str(value).splitlines()), default=0)
        )
    return float(min(max(longest + 2, 12), 42))


def _add_excel_table(worksheet, *, name: str) -> None:
    if worksheet.max_row < 2 or worksheet.max_column < 1:
        worksheet.auto_filter.ref = worksheet.dimensions
        return
    table = Table(displayName=name, ref=worksheet.dimensions)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    worksheet.add_table(table)


def _localized_labels(language: str, *, comparison_mode: bool) -> dict[str, str]:
    if language == "ru":
        return {
            "summary": "Сводка",
            "comparison": "Сравнение" if comparison_mode else "Данные",
            "sources": "Источники",
            "generated": "Создано",
            "instructions": "Что сравнить" if comparison_mode else "Цель",
            "currency": "Валюта",
            "documents": "Файлов",
            "rows": "Строк данных",
            "document": "Файл",
            "sheet": "Лист",
            "source_row": "Строка в источнике",
            "output_row": (
                "Строка в сравнении" if comparison_mode else "Строка в результате"
            ),
            "data_rows": "Строк данных",
            "columns": "Столбцов",
            "result": "Готовая таблица",
            "result_hint": "Данные приведены к единому виду. Исходники перечислены на отдельном листе.",
        }
    return {
        "summary": "Summary",
        "comparison": "Comparison" if comparison_mode else "Data",
        "sources": "Sources",
        "generated": "Generated",
        "instructions": "Comparison instructions" if comparison_mode else "Goal",
        "currency": "Currency",
        "documents": "Documents",
        "rows": "Data rows",
        "document": "Document",
        "sheet": "Sheet",
        "source_row": "Source row",
        "output_row": "Comparison row" if comparison_mode else "Output row",
        "data_rows": "Data rows",
        "columns": "Columns",
        "result": "Completed spreadsheet",
        "result_hint": "The data is normalized into one table. Source evidence is kept on a separate sheet.",
    }


def render_comparison_workbook(
    *,
    tables: tuple[SourceTable, ...],
    target_path: Path,
    language: str,
    currency: str | None,
    instructions: str | None,
    column_schema: ComparisonColumnSchema | None = None,
    comparison_mode: bool = True,
) -> RenderedComparison:
    if not tables:
        raise ComparisonInputError("at least one source table is required")
    total_rows = sum(len(table.rows) for table in tables)
    if total_rows > _MAX_OUTPUT_ROWS:
        raise ComparisonInputError("combined workbook contains too many data rows")

    labels = _localized_labels(language, comparison_mode=comparison_mode)
    ordered_headers: list[str] = []
    if column_schema is None:
        canonical_headers: dict[str, str] = {}
        for table in tables:
            for header in table.headers:
                key = header.casefold()
                if key not in canonical_headers:
                    canonical_headers[key] = header
                    ordered_headers.append(header)
    else:
        ordered_headers.extend(column_schema.canonical_headers)
        seen_headers: set[str] = set()
        for header in ordered_headers:
            normalized = header.casefold()
            if not header.strip() or normalized in seen_headers:
                raise ComparisonInputError(
                    "comparison schema contains invalid canonical columns"
                )
            seen_headers.add(normalized)
        for table in tables:
            for source_header in table.headers:
                canonical_header = column_schema.canonical_header(table, source_header)
                if canonical_header.casefold() not in seen_headers:
                    raise ComparisonInputError(
                        "comparison schema references an unknown canonical column"
                    )

    normalized_rows: list[list[Any]] = []
    source_records: list[ComparisonSource] = []
    output_row = 2
    for table in tables:
        if column_schema is None:
            header_positions = {
                header.casefold(): index for index, header in enumerate(table.headers)
            }
        else:
            header_positions: dict[str, int] = {}
            for index, header in enumerate(table.headers):
                canonical_key = column_schema.canonical_header(table, header).casefold()
                if canonical_key in header_positions:
                    raise ComparisonInputError(
                        "comparison schema merges columns within one source table"
                    )
                header_positions[canonical_key] = index
        first_output_row = output_row
        for row in table.rows:
            normalized_rows.append(
                [
                    row[position]
                    if position is not None and position < len(row)
                    else None
                    for header in ordered_headers
                    for position in [header_positions.get(header.casefold())]
                ]
            )
            output_row += 1
        if table.rows:
            source_records.append(
                ComparisonSource(
                    document_id=table.document_id,
                    filename=table.filename,
                    sheet_name=table.sheet_name,
                    row_start=table.first_data_row,
                    row_end=table.first_data_row + len(table.rows) - 1,
                    output_row_start=first_output_row,
                    output_row_end=output_row - 1,
                )
            )

    data_types = [
        _infer_data_type(row[index] for row in normalized_rows)
        for index in range(len(ordered_headers))
    ]
    normalized_rows = [
        [_coerce_cell(value, data_types[index]) for index, value in enumerate(row)]
        for row in normalized_rows
    ]

    workbook = Workbook()
    summary = workbook.active
    summary.title = labels["summary"]
    comparison = workbook.create_sheet(labels["comparison"])
    sources_sheet = workbook.create_sheet(labels["sources"])

    accent = "5B5BD6"
    accent_dark = "4545B8"
    foreground = "25252D"
    border_color = "DADCE5"
    header_fill = PatternFill("solid", fgColor=accent)
    header_font = Font(color="FFFFFF", bold=True)
    subtle_fill = PatternFill("solid", fgColor="F2F3F8")
    section_border = Border(bottom=Side(style="thin", color=border_color))

    for worksheet in (summary, comparison, sources_sheet):
        worksheet.sheet_view.showGridLines = False

    summary.merge_cells("A1:B1")
    summary["A1"] = labels["result"]
    summary["A1"].fill = PatternFill("solid", fgColor=accent_dark)
    summary["A1"].font = Font(color="FFFFFF", bold=True, size=18)
    summary["A1"].alignment = Alignment(vertical="center")
    summary.row_dimensions[1].height = 34
    summary.merge_cells("A2:B2")
    summary["A2"] = labels["result_hint"]
    summary["A2"].fill = subtle_fill
    summary["A2"].font = Font(color="5F6170", italic=True)
    summary["A2"].alignment = Alignment(wrap_text=True, vertical="center")
    summary.row_dimensions[2].height = 34

    summary_rows = [
        (
            labels["generated"],
            datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC"),
        ),
        (labels["instructions"], instructions or ""),
        (labels["currency"], currency or ""),
        (labels["documents"], len({table.document_id for table in tables})),
        (labels["rows"], total_rows),
    ]
    summary.append((None, None))
    for row in summary_rows:
        summary.append(row)
    for row in summary.iter_rows(min_row=4, max_col=2):
        row[0].font = Font(bold=True, color=foreground)
        row[0].fill = subtle_fill
        row[0].border = section_border
        row[1].border = section_border
        row[0].alignment = Alignment(vertical="top", wrap_text=True)
        row[1].alignment = Alignment(vertical="top", wrap_text=True)
    summary.row_dimensions[5].height = 46
    summary.column_dimensions["A"].width = 26
    summary.column_dimensions["B"].width = 68

    comparison.append([_safe_cell(value) for value in ordered_headers])
    for cell in comparison[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    for row in normalized_rows:
        comparison.append(row)
    comparison.freeze_panes = "A2"
    comparison.row_dimensions[1].height = 30
    for index, (header, data_type) in enumerate(
        zip(ordered_headers, data_types, strict=True), start=1
    ):
        values = [row[index - 1] for row in normalized_rows]
        comparison.column_dimensions[get_column_letter(index)].width = _column_width(
            header, values
        )
        number_format = _column_format(header, data_type, currency)
        for cell in comparison[get_column_letter(index)][1:]:
            cell.alignment = Alignment(
                horizontal="right" if data_type == "number" else "left",
                vertical="top",
                wrap_text=data_type == "text",
            )
            if number_format:
                cell.number_format = number_format
    _add_excel_table(comparison, name="LightnyData")

    source_headers = [
        labels["document"],
        labels["sheet"],
        labels["data_rows"],
        labels["columns"],
        labels["source_row"],
        labels["output_row"],
    ]
    sources_sheet.append(source_headers)
    for cell in sources_sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
    for table, source in zip(
        (table for table in tables if table.rows), source_records, strict=True
    ):
        sources_sheet.append(
            [
                _safe_cell(table.filename),
                _safe_cell(table.sheet_name),
                len(table.rows),
                len(table.headers),
                f"{source.row_start}-{source.row_end}",
                f"{source.output_row_start}-{source.output_row_end}",
            ]
        )
    sources_sheet.freeze_panes = "A2"
    sources_sheet.row_dimensions[1].height = 30
    for index, width in enumerate((36, 22, 16, 14, 22, 24), start=1):
        sources_sheet.column_dimensions[get_column_letter(index)].width = width
    for row in sources_sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    _add_excel_table(sources_sheet, name="LightnySources")

    preview_column_count = min(len(ordered_headers), _PREVIEW_MAX_COLUMNS)
    preview_row_count = min(len(normalized_rows), _PREVIEW_MAX_ROWS)
    warning_codes: list[str] = []
    if len(normalized_rows) > preview_row_count:
        warning_codes.append("preview_rows_truncated")
    if len(ordered_headers) > preview_column_count:
        warning_codes.append("preview_columns_truncated")
    if any(
        isinstance(value, str) and len(value) > _PREVIEW_MAX_CELL_CHARS
        for row in normalized_rows[:preview_row_count]
        for value in row[:preview_column_count]
    ):
        warning_codes.append("preview_cells_truncated")
    preview = {
        "version": 1,
        "goal": instructions,
        "row_count": total_rows,
        "column_count": len(ordered_headers),
        "source_count": len(source_records),
        "columns": [
            {
                "label": header,
                "data_type": data_types[index],
                "number_format": _column_format(header, data_types[index], currency),
            }
            for index, header in enumerate(ordered_headers[:preview_column_count])
        ],
        "rows": [
            [_preview_value(value) for value in row[:preview_column_count]]
            for row in normalized_rows[:preview_row_count]
        ],
        "rows_truncated": total_rows > preview_row_count,
        "columns_truncated": len(ordered_headers) > preview_column_count,
        "warning_codes": warning_codes,
    }
    while (
        preview["rows"]
        and len(
            json.dumps(preview, ensure_ascii=False, separators=(",", ":")).encode(
                "utf-8"
            )
        )
        > _PREVIEW_MAX_JSON_BYTES
    ):
        preview["rows"].pop()
        preview["rows_truncated"] = True
        if "preview_rows_truncated" not in warning_codes:
            warning_codes.insert(0, "preview_rows_truncated")

    workbook.save(target_path)
    workbook.close()
    return RenderedComparison(
        path=target_path,
        row_count=total_rows,
        column_count=len(ordered_headers),
        sources=tuple(source_records),
        preview=preview,
    )


def validate_rendered_workbook(path: Path) -> None:
    inspect_spreadsheet_source(str(path), path.name)
    workbook = load_workbook(path, read_only=False, data_only=False)
    try:
        if len(workbook.sheetnames) != 3:
            raise ComparisonInputError("rendered workbook must contain three sheets")
        if workbook[workbook.sheetnames[1]].max_row < 1:
            raise ComparisonInputError("rendered workbook has no header row")
        data_sheet = workbook[workbook.sheetnames[1]]
        sources_sheet = workbook[workbook.sheetnames[2]]
        if data_sheet.max_row >= 2 and "LightnyData" not in data_sheet.tables:
            raise ComparisonInputError("rendered workbook data must use an Excel table")
        if sources_sheet.max_row >= 2 and "LightnySources" not in sources_sheet.tables:
            raise ComparisonInputError(
                "rendered workbook sources must use an Excel table"
            )
    finally:
        workbook.close()
