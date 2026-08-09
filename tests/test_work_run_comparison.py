from __future__ import annotations

import json
import uuid

from openpyxl import load_workbook

from app.services.work_runs.comparison import (
    ComparisonColumnSchema,
    SourceTable,
    load_source_tables,
    render_comparison_workbook,
    validate_rendered_workbook,
)


def test_comparison_renderer_merges_columns_and_preserves_sources(tmp_path) -> None:
    first_id = uuid.uuid4()
    second_id = uuid.uuid4()
    first_path = tmp_path / "offer-a.csv"
    second_path = tmp_path / "offer-b.csv"
    first_path.write_text(
        'Товар;Цена;Комментарий\nЧай;120;=HYPERLINK("https://bad")\nКофе;350;Опт\n',
        encoding="utf-8",
    )
    second_path.write_text(
        "товар;Цена;Срок поставки\nЧай;110;3 дня\n",
        encoding="utf-8",
    )

    tables = (
        *load_source_tables(
            document_id=first_id, filename="offer-a.csv", path=first_path
        ),
        *load_source_tables(
            document_id=second_id, filename="offer-b.csv", path=second_path
        ),
    )
    target_path = tmp_path / "offer-comparison.xlsx"
    rendered = render_comparison_workbook(
        tables=tables,
        target_path=target_path,
        language="ru",
        currency="RUB",
        instructions="Сравнить цены и сроки поставки",
    )
    validate_rendered_workbook(target_path)

    assert rendered.row_count == 3
    assert rendered.column_count == 4
    assert len(rendered.sources) == 2
    assert rendered.preview["columns"] == [
        {"label": "Товар", "data_type": "text", "number_format": None},
        {
            "label": "Цена",
            "data_type": "number",
            "number_format": '#,##0.00 "RUB"',
        },
        {"label": "Комментарий", "data_type": "text", "number_format": None},
        {
            "label": "Срок поставки",
            "data_type": "text",
            "number_format": None,
        },
    ]
    assert rendered.preview["rows"][0] == [
        "Чай",
        120,
        '=HYPERLINK("https://bad")',
        None,
    ]
    workbook = load_workbook(target_path, data_only=False)
    try:
        assert workbook.sheetnames == ["Сводка", "Сравнение", "Источники"]
        comparison = workbook["Сравнение"]
        headers = [cell.value for cell in comparison[1]]
        assert headers == ["Товар", "Цена", "Комментарий", "Срок поставки"]
        assert comparison[2][2].value == '\'=HYPERLINK("https://bad")'
        assert comparison[2][1].value == 120
        assert comparison[2][1].number_format == '#,##0.00 "RUB"'
        assert "LightnyData" in comparison.tables
        sources = workbook["Источники"]
        assert sources.max_row == 3
        assert "LightnySources" in sources.tables
        assert sources.column_dimensions["A"].width == 36
        assert sources["A2"].alignment.wrap_text is True
    finally:
        workbook.close()


def test_comparison_reader_accepts_windows_1251_csv(tmp_path) -> None:
    path = tmp_path / "price.csv"
    path.write_bytes("Товар;Цена\nЧай;100\n".encode("cp1251"))

    tables = load_source_tables(
        document_id=uuid.uuid4(), filename="price.csv", path=path
    )

    assert tables[0].headers == ("Товар", "Цена")
    assert tables[0].rows == (("Чай", "100"),)


def test_comparison_renderer_applies_validated_column_schema(tmp_path) -> None:
    first_path = tmp_path / "offer-a.csv"
    second_path = tmp_path / "offer-b.csv"
    first_path.write_text("Наименование;Цена\nЧай;120\n", encoding="utf-8")
    second_path.write_text("Товар;Стоимость\nЧай;110\n", encoding="utf-8")
    first = load_source_tables(
        document_id=uuid.uuid4(), filename="offer-a.csv", path=first_path
    )[0]
    second = load_source_tables(
        document_id=uuid.uuid4(), filename="offer-b.csv", path=second_path
    )[0]
    schema = ComparisonColumnSchema(
        canonical_headers=("Товар", "Цена"),
        source_headers={
            (first.document_id, first.sheet_name, "Наименование"): "Товар",
            (first.document_id, first.sheet_name, "Цена"): "Цена",
            (second.document_id, second.sheet_name, "Товар"): "Товар",
            (second.document_id, second.sheet_name, "Стоимость"): "Цена",
        },
    )
    target_path = tmp_path / "normalized.xlsx"

    render_comparison_workbook(
        tables=(first, second),
        target_path=target_path,
        language="ru",
        currency="RUB",
        instructions=None,
        column_schema=schema,
    )

    workbook = load_workbook(target_path, data_only=False)
    try:
        comparison = workbook["Сравнение"]
        assert [cell.value for cell in comparison[1]] == ["Товар", "Цена"]
        assert comparison[2][0].value == "Чай"
        assert comparison[2][1].value == 120
        assert comparison[3][0].value == "Чай"
        assert comparison[3][1].value == 110
    finally:
        workbook.close()


def test_generalized_renderer_uses_data_and_goal_labels(tmp_path) -> None:
    source_path = tmp_path / "inventory.csv"
    source_path.write_text("Item;Stock\nTea;12\n", encoding="utf-8")
    table = load_source_tables(
        document_id=uuid.uuid4(), filename="inventory.csv", path=source_path
    )[0]
    target_path = tmp_path / "spreadsheet.xlsx"

    render_comparison_workbook(
        tables=(table,),
        target_path=target_path,
        language="en",
        currency=None,
        instructions="Build a clean inventory table",
        comparison_mode=False,
    )

    workbook = load_workbook(target_path, data_only=False)
    try:
        assert workbook.sheetnames == ["Summary", "Data", "Sources"]
        assert workbook["Summary"]["A5"].value == "Goal"
        assert workbook["Summary"]["B5"].value == "Build a clean inventory table"
        assert workbook["Sources"]["F1"].value == "Output row"
        assert "LightnyData" in workbook["Data"].tables
    finally:
        workbook.close()


def test_preview_stays_within_private_transport_limit(tmp_path) -> None:
    table = SourceTable(
        document_id=uuid.uuid4(),
        filename="wide.xlsx",
        sheet_name="Data",
        headers=tuple(f"Column {index}" for index in range(30)),
        rows=tuple(tuple("🙂" * 500 for _ in range(30)) for _ in range(40)),
        first_data_row=2,
    )

    rendered = render_comparison_workbook(
        tables=(table,),
        target_path=tmp_path / "wide.xlsx",
        language="en",
        currency=None,
        instructions="Keep this readable",
        comparison_mode=False,
    )

    encoded = json.dumps(
        rendered.preview,
        ensure_ascii=False,
        separators=(",", ":"),
    ).encode("utf-8")
    assert len(encoded) <= 1_800_000
    assert rendered.preview["rows_truncated"] is True
    assert "preview_rows_truncated" in rendered.preview["warning_codes"]
