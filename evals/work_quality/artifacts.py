from __future__ import annotations

import csv
import io
import zipfile
from pathlib import Path

from openpyxl import load_workbook


def validate_artifact_file(path: Path) -> tuple[bool, str]:
    if not path.is_file():
        return False, "downloaded artifact is missing"
    if path.stat().st_size == 0:
        return False, "downloaded artifact is empty"

    extension = path.suffix.lower()
    try:
        if extension == ".xlsx":
            workbook = load_workbook(path, read_only=True, data_only=False)
            try:
                if not workbook.sheetnames:
                    return False, "workbook has no worksheets"
                if not any(
                    sheet.max_row > 0 and sheet.max_column > 0
                    for sheet in workbook.worksheets
                ):
                    return False, "workbook has no populated cells"
            finally:
                workbook.close()
        elif extension == ".docx":
            _require_zip_members(path, {"[Content_Types].xml", "word/document.xml"})
        elif extension == ".pptx":
            _require_zip_members(path, {"[Content_Types].xml", "ppt/presentation.xml"})
        elif extension == ".pdf":
            payload = path.read_bytes()
            if not payload.startswith(b"%PDF-") or b"%%EOF" not in payload[-2048:]:
                return False, "PDF signature or trailer is invalid"
        elif extension == ".csv":
            payload = path.read_text(encoding="utf-8-sig")
            rows = list(csv.reader(io.StringIO(payload)))
            if len(rows) < 2 or not rows[0]:
                return False, "CSV has no header and data row"
        elif extension in {".md", ".txt"}:
            if not path.read_text(encoding="utf-8").strip():
                return False, "text artifact is blank"
        else:
            return (
                False,
                f"no structural validator for {extension or 'extensionless file'}",
            )
    except (OSError, UnicodeError, ValueError, zipfile.BadZipFile) as exc:
        return False, f"artifact could not be parsed: {exc}"
    return True, "artifact is structurally valid"


def _require_zip_members(path: Path, required: set[str]) -> None:
    with zipfile.ZipFile(path) as archive:
        missing = required - set(archive.namelist())
    if missing:
        raise ValueError(f"archive is missing required members: {sorted(missing)}")
